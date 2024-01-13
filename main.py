import telebot
import my_token
from openpyxl import load_workbook
from telebot import types

book = load_workbook(filename="Book.xlsx")
genre_sheet = book["Genres"]
user_sheet = book["User-Genre"]

bot = telebot.TeleBot(my_token.telebot_token)
bot.remove_webhook()

# user_genres = []
vacant_genres = []

current_user = None


@bot.message_handler(commands=['start'])
def start(message):

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)

    support = types.KeyboardButton("Поддержка")
    genres = types.KeyboardButton("Жанры")
    predict = types.KeyboardButton("Получить плейлист!")

    markup.add(support, genres, predict)

    bot.send_message(message.chat.id, f'Добрый день, {message.from_user.first_name}!\n'
                                      f'Этот бот создан для того, чтобы помочь Вам осознать собственное настроение и '
                                      f'на его основе составить музыкальный плейлист, который Вам точно понравится.\n'
                                      f'Вы можете начать работу с ним прямо сейчас или сначала настроить свои любимые '
                                      f'жанры, чтобы наша подборка понравилась Вам еще сильнее!', reply_markup=
                                      markup)
    bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAELKLZloTEXxO9gsgWaPcTka2f-xskchQAC-BYAAstG0UpJuy3W8rP2XDQE')


def add_user(xlsx_user_id, user_id):
    for i in range(15):
        if i == 0:
            user_sheet.cell(row=xlsx_user_id, column=i + 1).value = user_id
        else:
            user_sheet.cell(row=xlsx_user_id, column=i + 1).value = 0
    book.save('Book.xlsx')


def find_or_add_user(user_id):
    j = 0
    for row in user_sheet.iter_rows(max_col=1):
        j += 1
        for cell in row:

            if cell.value is None:
                print(f'current user = {j}, id = {user_id}')
                add_user(j, user_id)
                return j

            elif cell.value == user_id:
                print(f'current user = {j}, id = {user_id}')
                return j


def find_user(user_id):
    j = 0
    for row in user_sheet.iter_rows(max_col=1):
        j += 1
        for cell in row:
            if cell.value == user_id:
                print(f'current user = {j}, id = {user_id}')
                return j


def genres_cmd(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    support = types.KeyboardButton("Добавить")
    genres = types.KeyboardButton("Убрать")
    menu = types.KeyboardButton("Вернуться в меню")
    markup.add(support, genres, menu)

    user_id = message.from_user.id
    user = find_or_add_user(user_id)
    print(user)

    user_genres = get_user_genres(user)

    bot.send_message(message.chat.id, f'Вот список ваших любимых жанров: \n'
                                      f'{user_genres}\n'
                                      f'Что бы вы хотели с ними сделать?', reply_markup=markup)


def predict_cmd(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    menu = types.KeyboardButton("Вернуться в меню")
    markup.add(menu)

    bot.send_message(message.chat.id, f'Если Вы пришлете свою фотографию в этот диалог, наша нейронная сеть '
                                      f'сможет определить Ваше настроение и составить плейлист по настроению!',
                     reply_markup=markup)


def return_cmd(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)

    support = types.KeyboardButton("Поддержка")
    genres = types.KeyboardButton("Жанры")
    predict = types.KeyboardButton("Получить плейлист!")

    markup.add(support, genres, predict)

    bot.send_message(message.chat.id, f'Еще раз здравствуйте, {message.from_user.first_name}! 🩵\n'
                                      f'Хотите начать работу с ботом или настроить свои любимые жанры?',
                     reply_markup=markup)

    bot.send_sticker(message.chat.id, 'CAACAgIAAxkBAAELKNFloT5zfTrwuqGw95BKsyz3_ytmxQACHxcAAkuY0Eo26-gZoGwtPDQE')


def support_cmd(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    menu = types.KeyboardButton("Вернуться в меню")
    markup.add(menu)

    bot.send_message(message.chat.id, f'Бот был создан совместно студентами третьего курса Университета ИТМО '
                                      f'Пряничниковым К. С. и Серебренниковой В. В.!🕊 \n'
                                      f'По любому вопросу можете обратиться сюда: @needlessbeating',
                     reply_markup=markup)


def add_cmd(message):
    all_genres = ""
    for i in range(13):
        c = genre_sheet.cell(row=i+1, column=1)
        v = c.value
        all_genres += v
        if i < 12:
            all_genres += ", "

    bot.send_message(message.chat.id, f'Следующие жанры учтены в боте: {all_genres}.\n'
                                      f'Чтобы добавить жанр, отправьте его название в следующем сообщении:')

    bot.register_next_step_handler(message, pick_genre)


def find_genre(message):
    i = 1
    for column in genre_sheet.iter_cols(1):
        for cell in column:
            i += 1
            if cell.value == message.text:
                return i


def find_first_vacant():
    j = 1
    for _ in user_sheet["A"]:
        j += 1
    print(j)
    return j


def pick_genre(message):
    try:
        user_id = message.from_user.id
        genre = find_genre(message)
        user = find_user(user_id)
        print(f'user = {user}, genre = {genre}')
        user_sheet.cell(row=user, column=genre).value = 1
        book.save('Book.xlsx')

    except Exception as e:
        print(e)
    bot.send_message(message.chat.id, "Успех!")


@bot.message_handler(content_types='text')
def message_reply(message):
    if message.text == "Жанры":
        genres_cmd(message)

    elif message.text == "Получить плейлист!":
        predict_cmd(message)

    elif message.text == "Вернуться на главную страницу":
        return_cmd(message)

    elif message.text == "Поддержка":
        support_cmd(message)

    elif message.text == "Добавить":
        add_cmd(message)


def get_user_genres(user_id):
    user_genres = []
    user_genres_names = ""
    for cell in user_sheet[user_id]:
        user_genres.append(cell.value)

    for i in range(len(user_genres)):
        if i == 0 or user_genres[i] == 0:
            continue
        else:
            if i == 14:
                user_genres_names += user_sheet.cell(2, i + 1).value
                user_genres_names += ", "
            else:
                user_genres_names += user_sheet.cell(2, i+1).value

    print(user_genres)
    print(user_genres_names)
    return user_genres_names


def get_all_genres(sheet):
    col = sheet.col_values(0)
    return col


'''

def pick_time(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    back = types.KeyboardButton("На главную страницу")
    markup.add(back)
    book = xlrd.open_workbook('Даты.xls', formatting_info=True)
    if message.text == "На главную страницу":
        menu(message)
    elif message.text == "Понедельник":
        sheet = book.sheet_by_index(0)
    elif message.text == "Вторник":
        sheet = book.sheet_by_index(1)
    elif message.text == "Среда":
        sheet = book.sheet_by_index(2)
    elif message.text == "Четверг":
        sheet = book.sheet_by_index(3)
    elif message.text == "Пятница":
        sheet = book.sheet_by_index(4)
    else:
        bot.send_message(message.chat.id, "Что-то пошло не так. День недели, который Вы ввели, недоступен.",
                                          reply_markup=markup)
        date(message)
    weekday.clear()
    weekday.insert(0, message.text)
    row = sheet.row_values(0)
    all_times.append(row)
    bot.send_message(message.chat.id, "Данное время доступно. Введите выбранное время в формате HH:MM, чтобы "
                                      "продолжить.")
    bot.send_message(message.chat.id, row, reply_markup=markup)
    bot.register_next_step_handler(message, time)


def time(message):
    c = bool()
    chosen_time.clear()
    chosen_time.append(message.text)
    result = []
    for element in all_times:
        if type(element) is list:
            result += element
        else:
            result.append(element)
    all_times_string = ''.join(result)
    chosen_time_string = ''.join(chosen_time)
    if all_times_string.find(chosen_time_string) != -1:
        c = True
    all_times.clear()
    if c:
        bot.send_message(message.chat.id, "Выбранное время: " + message.text)
        mail_check(message)
    else:
        bot.send_message(message.chat.id, "Данное время недоступно, либо то, что вы ввели, не является временем.")
        date(message)


def num(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    back = types.KeyboardButton("На главную страницу")
    markup.add(back)
    bot.send_message(message.chat.id, "Теперь введите номер заказа, совершенный с почты " + ''.join(mail) + ".",
                     reply_markup=markup)
    if message.text == 'На главную страницу':
        menu(message)
    else:
        bot.register_next_step_handler(message, regex_num)


def regex_num(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    back = types.KeyboardButton("На главную страницу")
    markup.add(back)
    if message.text == 'На главную страницу':
        bot.register_next_step_handler(message, menu)
    elif re.compile(r'\d{6}').match(message.text):
        number.clear()
        number.insert(0, message.text)
        bot.send_message(message.chat.id, "Спасибо!")
        validation(message)
    else:
        bot.send_message(message.chat.id, "Упс :( Указанный номер заказа не валиден.", reply_markup=markup)
        num(message)


def validation(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    yes = types.KeyboardButton("Верно")
    no = types.KeyboardButton("Неверно")
    markup.add(yes, no)
    if message.text == 'На главную страницу':
        menu(message)
    bot.send_message(message.chat.id, "Теперь, пожалуйста, проверьте введенные данные:\nПочта: " + ''.join(mail) +
                     "\nНомер заказа :" + ''.join(number), reply_markup=markup)
    bot.register_next_step_handler(message, yes_or_no)


def yes_or_no(message):
    if message.text == "Верно":
        signup = "\n" + ''.join(weekday) + "\n" + ''.join(chosen_time) + "\n" + ''.join(mail) + "\n" + ''.join(number) \
                 + "\n"
        f = open("signup.txt", "a")
        f.write(signup)
        f.close()
        bot.send_message(message.chat.id, "Спасибо за бронирование времени! Как только оператор одобрит Ваш заказ, "
                                          "курьер приедет к Вам!")
        menu(message)
    elif message.text == "Неверно":
        mail.clear()
        number.clear()
        all_times.clear()
        chosen_time.clear()
        weekday.clear()
        bot.send_message(message.chat.id, "Придётся начать сначала(")
        date(message)


def menu(message):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    support = types.KeyboardButton("Поддержка")
    pick_a_date = types.KeyboardButton("Выбрать дату")
    info = types.KeyboardButton("Информация")
    markup.add(support, pick_a_date, info)
    photo = open("whatnext.jpg", 'rb')
    bot.send_photo(message.chat.id, photo)
    bot.send_message(message.chat.id, f'Добрый день, {message.from_user.first_name}! Что бы Вы хотели сделать?',
                     reply_markup=markup)
    bot.register_next_step_handler(message, menu_jumper)


def menu_jumper(message):
    if message.text == "Выбрать дату":
        date(message)
    elif message.text == "Информация":
        information(message)
    elif message.text == "Поддержка":
        helper(message)
    else:
        bot.send_message(message.chat.id, "Возникла ошибка. Такой команды нет в меню, попробуйте ещё раз.")
        menu(message)

'''
if __name__ == '__main__':
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(e)
